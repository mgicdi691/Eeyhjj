import os, io, sqlite3, logging, asyncio
from datetime import datetime
from pathlib import Path
from html import escape

from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

BOT_TOKEN = os.getenv('BOT_TOKEN', '8675144844:AAH6j5flV7e_Yn3OrCBS6dcUoDpzEleajvs')
ADMIN_IDS = {int(x) for x in os.getenv('ADMIN_IDS', '6995426618').split(',') if x.strip().isdigit()}
DB_FILE = os.getenv('DB_FILE', 'store.db')
UPLOAD_DIR = Path(os.getenv('UPLOAD_DIR', 'product_files')); UPLOAD_DIR.mkdir(exist_ok=True)
TEMP_NUMBER_URL = 'https://t.me/OtpNowBoost_bot'
MAIL_INBOX_URL = 'https://dongvanfb.net/get_code_mail/'
# Put valid Telegram custom-emoji IDs here. Empty values intentionally fall back to text labels, never fake IDs.
CUSTOM_EMOJI_IDS = {
    'home': os.getenv('CE_HOME',''), 'success': os.getenv('CE_SUCCESS',''), 'error': os.getenv('CE_ERROR',''),
    'money': os.getenv('CE_MONEY',''), 'product': os.getenv('CE_PRODUCT',''), 'admin': os.getenv('CE_ADMIN',''),
    'join': os.getenv('CE_JOIN',''), 'payment': os.getenv('CE_PAYMENT',''), 'settings': os.getenv('CE_SETTINGS','')
}
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(name)s: %(message)s')
log=logging.getLogger('storebot')

def now(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
def db():
    c=sqlite3.connect(DB_FILE, timeout=30); c.row_factory=sqlite3.Row; return c
def money(v): return f'{float(v):.2f}'.rstrip('0').rstrip('.')
def is_admin(uid): return uid in ADMIN_IDS

def ce(kind):
    eid=CUSTOM_EMOJI_IDS.get(kind,'')
    return f'<tg-emoji emoji-id="{escape(eid)}"></tg-emoji>' if eid else ''

def msg(kind,title,body=''):
    icon=ce(kind)
    return f'{icon}<b>{escape(title)}</b>\n\n{body}' if icon else f'<b>{escape(title)}</b>\n\n{body}'

def init_db():
    c=db(); q=c.execute
    q('''CREATE TABLE IF NOT EXISTS users(user_id INTEGER PRIMARY KEY, username TEXT DEFAULT '', first_name TEXT DEFAULT '', balance REAL DEFAULT 0, created_at TEXT NOT NULL)''')
    q('''CREATE TABLE IF NOT EXISTS products(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL,price REAL NOT NULL,description TEXT DEFAULT '',active INTEGER DEFAULT 1,created_at TEXT NOT NULL)''')
    q('''CREATE TABLE IF NOT EXISTS stock(id INTEGER PRIMARY KEY AUTOINCREMENT,product_id INTEGER NOT NULL,file_id TEXT NOT NULL,file_name TEXT DEFAULT '',sold INTEGER DEFAULT 0,sold_to INTEGER,sold_at TEXT,FOREIGN KEY(product_id) REFERENCES products(id))''')
    q('''CREATE TABLE IF NOT EXISTS orders(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER NOT NULL,product_id INTEGER NOT NULL,quantity INTEGER NOT NULL,total REAL NOT NULL,status TEXT NOT NULL,created_at TEXT NOT NULL)''')
    q('''CREATE TABLE IF NOT EXISTS deposits(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER NOT NULL,amount REAL NOT NULL,trx_id TEXT DEFAULT '',status TEXT NOT NULL,created_at TEXT NOT NULL)''')
    q('''CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT DEFAULT '')''')
    q('''CREATE TABLE IF NOT EXISTS payment_categories(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE NOT NULL,details TEXT DEFAULT '',enabled INTEGER DEFAULT 1,min_deposit REAL DEFAULT 0,max_deposit REAL DEFAULT 0)''')
    q('''CREATE TABLE IF NOT EXISTS force_join(id INTEGER PRIMARY KEY AUTOINCREMENT,chat TEXT UNIQUE NOT NULL,enabled INTEGER DEFAULT 1)''')
    c.commit(); c.close()

def setting(k,default=''):
    c=db(); r=c.execute('SELECT value FROM settings WHERE key=?',(k,)).fetchone(); c.close(); return r['value'] if r else default
def set_setting(k,v):
    c=db(); c.execute('INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',(k,str(v))); c.commit(); c.close()

def ensure_user(u):
    c=db(); c.execute('''INSERT INTO users(user_id,username,first_name,created_at) VALUES(?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET username=excluded.username,first_name=excluded.first_name''',(u.id,u.username or '',u.first_name or '',now())); c.commit(); c.close()
def balance(uid):
    c=db(); r=c.execute('SELECT balance FROM users WHERE user_id=?',(uid,)).fetchone(); c.close(); return float(r['balance']) if r else 0.0

def products():
    c=db(); rows=c.execute('''SELECT p.*,COALESCE((SELECT COUNT(*) FROM stock s WHERE s.product_id=p.id AND s.sold=0),0) stock_count FROM products p WHERE p.active=1 ORDER BY p.id DESC''').fetchall(); c.close(); return rows

def product(pid):
    c=db(); r=c.execute('SELECT * FROM products WHERE id=?',(pid,)).fetchone(); c.close(); return r

def user_kb(uid):
    return ReplyKeyboardMarkup([
        [KeyboardButton('Buy Product'),KeyboardButton('Deposit Money')],
        [KeyboardButton('My Balance'),KeyboardButton('Price List')],
        [KeyboardButton('Temp Number'),KeyboardButton('Mail Inbox')],
        *([[KeyboardButton('Admin Panel')]] if is_admin(uid) else [])
    ],resize_keyboard=True)

def admin_kb():
    return ReplyKeyboardMarkup([
        [KeyboardButton('Product Management'),KeyboardButton('Force Join')],
        [KeyboardButton('Payment Management'),KeyboardButton('User Management')],
        [KeyboardButton('Pending Deposits'),KeyboardButton('Broadcast')],
        [KeyboardButton('Back')]
    ],resize_keyboard=True)

def product_admin_kb():
    return ReplyKeyboardMarkup([[KeyboardButton('Add Product'),KeyboardButton('Edit Product')],[KeyboardButton('Delete Product'),KeyboardButton('Products')],[KeyboardButton('Add Stock'),KeyboardButton('Back')]],resize_keyboard=True)
def force_kb():
    return ReplyKeyboardMarkup([[KeyboardButton('Add Force Join'),KeyboardButton('Remove Force Join')],[KeyboardButton('Force Join List'),KeyboardButton('Force Join Active/Disable')],[KeyboardButton('Back')]],resize_keyboard=True)
def payment_kb():
    return ReplyKeyboardMarkup([[KeyboardButton('Add Category'),KeyboardButton('Edit Category')],[KeyboardButton('Delete Category'),KeyboardButton('Payment Categories')],[KeyboardButton('Payment Enable/Disable'),KeyboardButton('Set Deposit Min/Max')],[KeyboardButton('Back')]],resize_keyboard=True)
def user_kb_admin():
    return ReplyKeyboardMarkup([[KeyboardButton('Users'),KeyboardButton('Add Balance')],[KeyboardButton('Deduct Balance'),KeyboardButton('User Details')],[KeyboardButton('Back')]],resize_keyboard=True)
def back_kb(): return ReplyKeyboardMarkup([[KeyboardButton('Back')]],resize_keyboard=True)

def inline_link(label,url): return InlineKeyboardMarkup([[InlineKeyboardButton(label,url=url)]])

def state_clear(ctx):
    ctx.user_data.clear()

def active_payment_categories():
    c=db(); rows=c.execute('SELECT * FROM payment_categories WHERE enabled=1 ORDER BY id').fetchall(); c.close(); return rows

def force_rows(active_only=True):
    c=db(); sql='SELECT * FROM force_join'+(' WHERE enabled=1' if active_only else '')+' ORDER BY id'; rows=c.execute(sql).fetchall(); c.close(); return rows

async def force_join_ok(ctx,uid):
    rows=force_rows(True)
    for r in rows:
        chat=r['chat']
        try:
            m=await ctx.bot.get_chat_member(chat,uid)
            if m.status in ('left','kicked'): return False
        except Exception:
            return False
    return True

async def guard(update,ctx):
    uid=update.effective_user.id
    if is_admin(uid): return True
    if not force_rows(True): return True
    if await force_join_ok(ctx,uid): return True
    buttons=[]
    for r in force_rows(True):
        try:
            ch=await ctx.bot.get_chat(r['chat']); link=ch.invite_link
            if not link:
                if str(r['chat']).startswith('@'): link='https://t.me/'+str(r['chat'])[1:]
                else: link='https://t.me/'+str(r['chat']).lstrip('@')
            buttons.append([InlineKeyboardButton('Join',url=link)])
        except Exception: pass
    buttons.append([InlineKeyboardButton('I Joined - Check',callback_data='check_join')])
    await update.effective_message.reply_text(msg('join','Join Required','Please join all required channels/groups, then press Check.'),parse_mode='HTML',reply_markup=InlineKeyboardMarkup(buttons))
    return False

async def start(update,ctx):
    ensure_user(update.effective_user); state_clear(ctx)
    if not await guard(update,ctx): return
    await update.message.reply_text(msg('home','Welcome','Use the menu below to continue.'),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id))

async def check_join_cb(update,ctx):
    q=update.callback_query; await q.answer()
    if await force_join_ok(ctx,q.from_user.id):
        await q.message.reply_text(msg('success','Verified','You can now use the bot.'),parse_mode='HTML',reply_markup=user_kb(q.from_user.id))
    else: await q.answer('Not all required channels are joined.',show_alert=True)

async def show_buy(update,ctx):
    ps=products()
    if not ps: await update.message.reply_text(msg('product','No Products','There are no active products right now.'),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id)); return
    rows=[]; pair=[]
    for p in ps:
        pair.append(KeyboardButton(f'Product #{p["id"]}'))
        if len(pair)==2: rows.append(pair); pair=[]
    if pair: rows.append(pair)
    rows.append([KeyboardButton('Back')])
    ctx.user_data['product_buttons']={f'Product #{p["id"]}':p['id'] for p in ps}
    await update.message.reply_text(msg('product','Choose Product','Select a product.'),parse_mode='HTML',reply_markup=ReplyKeyboardMarkup(rows,resize_keyboard=True))

async def show_product(update,ctx,pid):
    p=product(pid)
    if not p or not p['active']: await update.message.reply_text(msg('error','Unavailable','Product is not available.'),parse_mode='HTML'); return
    c=db(); stock=c.execute('SELECT COUNT(*) c FROM stock WHERE product_id=? AND sold=0',(pid,)).fetchone()['c']; c.close()
    if stock<1: await update.message.reply_text(msg('error','Out of Stock','This product is currently out of stock.'),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id)); return
    ctx.user_data['buy_product_id']=pid
    body=f'<b>{escape(p["name"])}</b>\nPrice: <b>{money(p["price"])} ৳</b>\nStock: <b>{stock}</b>\n\n{escape(p["description"] or "No description.")}\n\nSend quantity:'
    await update.message.reply_text(msg('product','Product Details',body),parse_mode='HTML',reply_markup=back_kb())

async def buy_quantity(update,ctx,qty):
    pid=ctx.user_data.get('buy_product_id')
    if not pid: return False
    if qty<1: await update.message.reply_text(msg('error','Invalid Quantity','Enter a positive number.'),parse_mode='HTML'); return True
    c=db()
    try:
        c.execute('BEGIN IMMEDIATE')
        p=c.execute('SELECT * FROM products WHERE id=? AND active=1',(pid,)).fetchone()
        if not p: raise ValueError('Product unavailable')
        rows=c.execute('SELECT * FROM stock WHERE product_id=? AND sold=0 ORDER BY id LIMIT ?',(pid,qty)).fetchall()
        if len(rows)<qty: raise ValueError(f'Only {len(rows)} stock available')
        total=float(p['price'])*qty
        u=c.execute('SELECT balance FROM users WHERE user_id=?',(update.effective_user.id,)).fetchone()
        if not u or float(u['balance'])<total: raise PermissionError(f'Insufficient balance. Need {money(total)} ৳')
        c.execute('UPDATE users SET balance=balance-? WHERE user_id=?',(total,update.effective_user.id))
        for r in rows: c.execute('UPDATE stock SET sold=1,sold_to=?,sold_at=? WHERE id=? AND sold=0',(update.effective_user.id,now(),r['id']))
        c.execute('INSERT INTO orders(user_id,product_id,quantity,total,status,created_at) VALUES(?,?,?,?,?,?)',(update.effective_user.id,pid,qty,total,'completed',now()))
        c.commit()
    except PermissionError as e:
        c.rollback(); c.close(); await update.message.reply_text(msg('money','Insufficient Balance',escape(str(e))),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id)); return True
    except ValueError as e:
        c.rollback(); c.close(); await update.message.reply_text(msg('error','Purchase Failed',escape(str(e))),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id)); return True
    except Exception:
        c.rollback(); c.close(); log.exception('purchase'); await update.message.reply_text(msg('error','Purchase Failed','Please try again.'),parse_mode='HTML'); return True
    c.close(); ctx.user_data.pop('buy_product_id',None)
    await update.message.reply_text(msg('success','Purchase Successful',f'Product: <b>{escape(p["name"])}</b>\nQuantity: <b>{qty}</b>\nTotal: <b>{money(total)} ৳</b>\nNew Balance: <b>{money(balance(update.effective_user.id))} ৳</b>'),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id))
    for r in rows:
        try: await ctx.bot.send_document(update.effective_user.id,r['file_id'],caption=f'{p["name"]} - purchased')
        except Exception: await ctx.bot.send_message(update.effective_user.id,f'{p["name"]}\n{r["file_id"]}')
    return True

async def deposit_start(update,ctx):
    cats=active_payment_categories()
    if not cats: await update.message.reply_text(msg('payment','Payment Unavailable','No payment method is enabled.'),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id)); return
    ctx.user_data['state']='deposit_category'; ctx.user_data['deposit_categories']={f'Payment #{r["id"]}':r['id'] for r in cats}
    rows=[[KeyboardButton(f'Payment #{r["id"]}') for r in cats[i:i+2]] for i in range(0,len(cats),2)]; rows.append([KeyboardButton('Back')])
    await update.message.reply_text(msg('payment','Select Payment Method','Choose a payment category.'),parse_mode='HTML',reply_markup=ReplyKeyboardMarkup(rows,resize_keyboard=True))

async def payment_category(update,ctx,rid):
    c=db(); r=c.execute('SELECT * FROM payment_categories WHERE id=? AND enabled=1',(rid,)).fetchone(); c.close()
    if not r: await update.message.reply_text(msg('error','Unavailable','Payment category is disabled.')); return
    ctx.user_data.update(state='deposit_amount',payment_category_id=rid)
    lim=[]
    if r['min_deposit']>0: lim.append(f'Min: <b>{money(r["min_deposit"])} ৳</b>')
    if r['max_deposit']>0: lim.append(f'Max: <b>{money(r["max_deposit"])} ৳</b>')
    body=f'<b>{escape(r["name"])}</b>\n{escape(r["details"] or "")}\n\n'+' | '.join(lim)+'\n\nSend deposit amount.'
    await update.message.reply_text(msg('payment','Payment Details',body),parse_mode='HTML',reply_markup=back_kb())

async def deposit_amount(update,ctx,text):
    try: amount=float(text); assert amount>0
    except: await update.message.reply_text(msg('error','Invalid Amount','Enter a valid amount.')); return
    rid=ctx.user_data.get('payment_category_id'); c=db(); r=c.execute('SELECT * FROM payment_categories WHERE id=? AND enabled=1',(rid,)).fetchone(); c.close()
    if not r: await update.message.reply_text(msg('error','Payment Error','Category unavailable.')); return
    if r['min_deposit'] and amount<r['min_deposit'] or r['max_deposit'] and amount>r['max_deposit']:
        await update.message.reply_text(msg('error','Amount Not Allowed',f'Min: {money(r["min_deposit"])} ৳\nMax: {money(r["max_deposit"])} ৳')); return
    ctx.user_data.update(state='deposit_trx',deposit_amount=amount)
    await update.message.reply_text(msg('payment','Submit Transaction ID',f'Amount: <b>{money(amount)} ৳</b>\nSend your Transaction ID.'),parse_mode='HTML',reply_markup=back_kb())

async def deposit_trx(update,ctx,text):
    amount=ctx.user_data.get('deposit_amount'); rid=ctx.user_data.get('payment_category_id')
    if not amount: await update.message.reply_text(msg('error','Session Expired','Start deposit again.')); return
    c=db(); cur=c.execute('INSERT INTO deposits(user_id,amount,trx_id,status,created_at) VALUES(?,?,?,?,?)',(update.effective_user.id,amount,text,'pending',now())); did=cur.lastrowid; c.commit(); c.close()
    for aid in ADMIN_IDS:
        try: await ctx.bot.send_message(aid,msg('payment','New Deposit',f'User: <code>{update.effective_user.id}</code>\nAmount: <b>{money(amount)} ৳</b>\nTRX: <code>{escape(text)}</code>\nDeposit ID: <code>{did}</code>\n\nUse Pending Deposits in Admin Panel.'),parse_mode='HTML')
        except: pass
    state_clear(ctx); await update.message.reply_text(msg('success','Deposit Submitted','Your deposit is pending admin verification.'),parse_mode='HTML',reply_markup=user_kb(update.effective_user.id))

# ---------- Admin product ----------
async def admin_panel(update,ctx): await update.message.reply_text(msg('admin','Admin Panel','Choose a management section.'),parse_mode='HTML',reply_markup=admin_kb())
async def product_management(update,ctx): await update.message.reply_text(msg('product','Product Management','Manage products and stock.'),parse_mode='HTML',reply_markup=product_admin_kb())
async def add_product_start(update,ctx): ctx.user_data['admin_state']='add_name'; await update.message.reply_text('Send product name.',reply_markup=back_kb())
async def add_product_name(update,ctx): ctx.user_data['new_name']=update.message.text.strip(); ctx.user_data['admin_state']='add_price'; await update.message.reply_text('Send product price.')
async def add_product_price(update,ctx):
    try: v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Send a valid price.'); return
    ctx.user_data['new_price']=v; ctx.user_data['admin_state']='add_desc'; await update.message.reply_text('Send product description.')
async def add_product_desc(update,ctx):
    c=db(); cur=c.execute('INSERT INTO products(name,price,description,active,created_at) VALUES(?,?,?,?,?)',(ctx.user_data['new_name'],ctx.user_data['new_price'],update.message.text.strip(),1,now())); pid=cur.lastrowid; c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Product Added',f'Product ID: <b>{pid}</b>'),parse_mode='HTML',reply_markup=product_admin_kb())
async def admin_products(update,ctx):
    ps=products(); body='\n\n'.join([f'ID <b>{p["id"]}</b> | <b>{escape(p["name"])}</b> | {money(p["price"])} ৳ | Stock {p["stock_count"]}' for p in ps]) or 'No products.'; await update.message.reply_text(msg('product','Products',body),parse_mode='HTML',reply_markup=product_admin_kb())
async def delete_product_start(update,ctx): ctx.user_data['admin_state']='delete_product'; await update.message.reply_text('Send product ID to delete.',reply_markup=back_kb())
async def delete_product(update,ctx):
    try: pid=int(update.message.text)
    except: await update.message.reply_text('Invalid product ID.'); return
    c=db(); cur=c.execute('UPDATE products SET active=0 WHERE id=?',(pid,)); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Product Deleted','Product disabled successfully.'),parse_mode='HTML',reply_markup=product_admin_kb())
async def edit_product_start(update,ctx): ctx.user_data['admin_state']='edit_id'; await update.message.reply_text('Send product ID to edit.',reply_markup=back_kb())
async def edit_product_id(update,ctx):
    try: pid=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    if not product(pid): await update.message.reply_text('Product not found.'); return
    ctx.user_data.update(edit_id=pid,admin_state='edit_name'); await update.message.reply_text('Send new name.')
async def edit_product_name(update,ctx): ctx.user_data['edit_name_value']=update.message.text.strip(); ctx.user_data['admin_state']='edit_price'; await update.message.reply_text('Send new price.')
async def edit_product_price(update,ctx):
    try: v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Invalid price.'); return
    ctx.user_data['edit_price_value']=v; ctx.user_data['admin_state']='edit_desc'; await update.message.reply_text('Send new description.')
async def edit_product_desc(update,ctx):
    c=db(); c.execute('UPDATE products SET name=?,price=?,description=? WHERE id=?',(ctx.user_data['edit_name_value'],ctx.user_data['edit_price_value'],update.message.text.strip(),ctx.user_data['edit_id'])); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Product Updated','Changes saved.'),parse_mode='HTML',reply_markup=product_admin_kb())

async def add_stock_start(update,ctx):
    ps=products(); ctx.user_data['admin_state']='stock_product'; await update.message.reply_text('Send Product ID for stock upload. XLSX rows or individual files are supported.',reply_markup=back_kb())
async def stock_product(update,ctx):
    try: pid=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    if not product(pid): await update.message.reply_text('Product not found.'); return
    ctx.user_data.update(stock_pid=pid,admin_state='stock_file'); await update.message.reply_text('Send a document. XLSX: each non-empty row becomes one stock item. Other files: one file = one stock item.')
async def stock_file(update,ctx):
    pid=ctx.user_data.get('stock_pid'); doc=update.message.document
    if not doc: return
    if (doc.file_name or '').lower().endswith('.xlsx'):
        data=io.BytesIO(); tg=await doc.get_file(); await tg.download_to_memory(data); data.seek(0)
        try:
            from openpyxl import load_workbook
            wb=load_workbook(data,read_only=True,data_only=True); ws=wb.active; count=0; c=db()
            for row in ws.iter_rows(values_only=True):
                vals=[str(x).strip() for x in row if x is not None and str(x).strip()]
                if not vals: continue
                item=' | '.join(vals); c.execute('INSERT INTO stock(product_id,file_id,file_name) VALUES(?,?,?)',(pid,item,doc.file_name or 'xlsx')); count+=1
            c.commit(); c.close(); wb.close(); await update.message.reply_text(msg('success','XLSX Stock Added',f'Rows added: <b>{count}</b>\nEach row is one stock item.'),parse_mode='HTML',reply_markup=product_admin_kb())
        except Exception as e: log.exception('xlsx'); await update.message.reply_text(f'XLSX processing failed: {escape(str(e))}')
    else:
        c=db(); c.execute('INSERT INTO stock(product_id,file_id,file_name) VALUES(?,?,?)',(pid,doc.file_id,doc.file_name or '')); c.commit(); c.close(); await update.message.reply_text(msg('success','Stock Added','One stock item added. Send another file or Back.'),parse_mode='HTML',reply_markup=product_admin_kb())

# ---------- Force Join ----------
async def force_panel(update,ctx): await update.message.reply_text(msg('join','Force Join','Manage required channels/groups.'),parse_mode='HTML',reply_markup=force_kb())
async def force_add(update,ctx): ctx.user_data['admin_state']='fj_add'; await update.message.reply_text('Send channel/group @username or numeric chat ID. Bot must be a member/admin there.',reply_markup=back_kb())
async def force_add_save(update,ctx):
    chat=update.message.text.strip(); c=db()
    try: c.execute('INSERT INTO force_join(chat,enabled) VALUES(?,1)',(chat,)); c.commit(); ok=True
    except sqlite3.IntegrityError: ok=False
    c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Force Join Added','Added successfully.' if ok else 'Already exists.'),parse_mode='HTML',reply_markup=force_kb())
async def force_list(update,ctx):
    rows=force_rows(False); body='\n'.join([f'ID <b>{r["id"]}</b> | <code>{escape(r["chat"])}</code> | {"ACTIVE" if r["enabled"] else "DISABLED"}' for r in rows]) or 'No channels/groups.'; await update.message.reply_text(msg('join','Force Join List',body),parse_mode='HTML',reply_markup=force_kb())
async def force_remove(update,ctx): ctx.user_data['admin_state']='fj_remove'; await update.message.reply_text('Send Force Join ID to remove.',reply_markup=back_kb())
async def force_remove_save(update,ctx):
    try: i=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    c=db(); c.execute('DELETE FROM force_join WHERE id=?',(i,)); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Force Join Removed','Removed.'),parse_mode='HTML',reply_markup=force_kb())
async def force_toggle(update,ctx): ctx.user_data['admin_state']='fj_toggle'; await update.message.reply_text('Send Force Join ID to toggle Active/Disable.',reply_markup=back_kb())
async def force_toggle_save(update,ctx):
    try: i=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    c=db(); c.execute('UPDATE force_join SET enabled=CASE enabled WHEN 1 THEN 0 ELSE 1 END WHERE id=?',(i,)); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Force Join Updated','Status changed.'),parse_mode='HTML',reply_markup=force_kb())

# ---------- Payment ----------
async def payment_panel(update,ctx): await update.message.reply_text(msg('payment','Payment Management','Manage categories, details, limits and status.'),parse_mode='HTML',reply_markup=payment_kb())
async def add_cat(update,ctx): ctx.user_data['admin_state']='pc_name'; await update.message.reply_text('Send payment category name (e.g. bKash, Nagad).',reply_markup=back_kb())
async def pc_name(update,ctx): ctx.user_data['pc_name']=update.message.text.strip(); ctx.user_data['admin_state']='pc_details'; await update.message.reply_text('Send payment details/number.')
async def pc_details(update,ctx): ctx.user_data['pc_details']=update.message.text.strip(); ctx.user_data['admin_state']='pc_min'; await update.message.reply_text('Send minimum deposit (0 for none).')
async def pc_min(update,ctx):
    try: v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Invalid minimum.'); return
    ctx.user_data['pc_min']=v; ctx.user_data['admin_state']='pc_max'; await update.message.reply_text('Send maximum deposit (0 for none).')
async def pc_max(update,ctx):
    try: v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Invalid maximum.'); return
    c=db(); c.execute('INSERT INTO payment_categories(name,details,enabled,min_deposit,max_deposit) VALUES(?,?,?,?,?)',(ctx.user_data['pc_name'],ctx.user_data['pc_details'],1,ctx.user_data['pc_min'],v)); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Payment Category Added','Category is enabled.'),parse_mode='HTML',reply_markup=payment_kb())
async def payment_list(update,ctx):
    c=db(); rows=c.execute('SELECT * FROM payment_categories ORDER BY id').fetchall(); c.close(); body='\n\n'.join([f'ID <b>{r["id"]}</b> | <b>{escape(r["name"])}</b>\nStatus: {"ON" if r["enabled"] else "OFF"}\nMin: {money(r["min_deposit"])} | Max: {money(r["max_deposit"])}\nDetails: {escape(r["details"])}' for r in rows]) or 'No payment categories.'; await update.message.reply_text(msg('payment','Payment Categories',body),parse_mode='HTML',reply_markup=payment_kb())
async def payment_edit(update,ctx): ctx.user_data['admin_state']='pe_id'; await update.message.reply_text('Send payment category ID to edit.')
async def pe_id(update,ctx):
    try:i=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    c=db(); r=c.execute('SELECT * FROM payment_categories WHERE id=?',(i,)).fetchone(); c.close()
    if not r: await update.message.reply_text('Not found.'); return
    ctx.user_data.update(pe_id=i,admin_state='pe_details'); await update.message.reply_text('Send new details/number.')
async def pe_details(update,ctx): ctx.user_data['pe_details']=update.message.text.strip(); ctx.user_data['admin_state']='pe_min'; await update.message.reply_text('Send new minimum deposit.')
async def pe_min(update,ctx):
    try:v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Invalid minimum.'); return
    ctx.user_data['pe_min']=v; ctx.user_data['admin_state']='pe_max'; await update.message.reply_text('Send new maximum deposit (0 for none).')
async def pe_max(update,ctx):
    try:v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Invalid maximum.'); return
    c=db(); c.execute('UPDATE payment_categories SET details=?,min_deposit=?,max_deposit=? WHERE id=?',(ctx.user_data['pe_details'],ctx.user_data['pe_min'],v,ctx.user_data['pe_id'])); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Payment Updated','Details and limits saved.'),parse_mode='HTML',reply_markup=payment_kb())
async def payment_delete(update,ctx): ctx.user_data['admin_state']='pd_id'; await update.message.reply_text('Send payment category ID to delete.')
async def payment_delete_save(update,ctx):
    try:i=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    c=db(); c.execute('DELETE FROM payment_categories WHERE id=?',(i,)); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Payment Deleted','Category removed.'),parse_mode='HTML',reply_markup=payment_kb())
async def payment_toggle(update,ctx): ctx.user_data['admin_state']='pt_id'; await update.message.reply_text('Send payment category ID to toggle enable/disable.')
async def payment_toggle_save(update,ctx):
    try:i=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    c=db(); c.execute('UPDATE payment_categories SET enabled=CASE enabled WHEN 1 THEN 0 ELSE 1 END WHERE id=?',(i,)); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Payment Status Updated','Enabled/disabled.'),parse_mode='HTML',reply_markup=payment_kb())
async def payment_limits(update,ctx): ctx.user_data['admin_state']='pl_id'; await update.message.reply_text('Send category ID to change min/max.')
async def payment_limits_id(update,ctx):
    try:i=int(update.message.text)
    except: await update.message.reply_text('Invalid ID.'); return
    ctx.user_data.update(pl_id=i,admin_state='pl_min'); await update.message.reply_text('Send new minimum deposit.')
async def payment_limits_min(update,ctx):
    try:v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Invalid minimum.'); return
    ctx.user_data['pl_min']=v; ctx.user_data['admin_state']='pl_max'; await update.message.reply_text('Send new maximum deposit.')
async def payment_limits_max(update,ctx):
    try:v=float(update.message.text); assert v>=0
    except: await update.message.reply_text('Invalid maximum.'); return
    c=db(); c.execute('UPDATE payment_categories SET min_deposit=?,max_deposit=? WHERE id=?',(ctx.user_data['pl_min'],v,ctx.user_data['pl_id'])); c.commit(); c.close(); state_clear(ctx); await update.message.reply_text(msg('success','Limits Updated','Deposit range saved.'),parse_mode='HTML',reply_markup=payment_kb())

# ---------- Users / deposits ----------
async def users(update,ctx):
    c=db(); rows=c.execute('SELECT user_id,username,first_name,balance FROM users ORDER BY user_id DESC LIMIT 50').fetchall(); c.close(); body='\n'.join([f'<code>{r["user_id"]}</code> | @{escape(r["username"] or "-")} | <b>{money(r["balance"])} ৳</b>' for r in rows]) or 'No users.'; await update.message.reply_text(msg('admin','Users',body),parse_mode='HTML',reply_markup=user_kb_admin())
async def balance_change_start(update,ctx,add=True): ctx.user_data['admin_state']='bal_uid_add' if add else 'bal_uid_sub'; await update.message.reply_text('Send user ID.')
async def balance_uid(update,ctx):
    try:uid=int(update.message.text)
    except: await update.message.reply_text('Invalid user ID.'); return
    if not (db().execute('SELECT 1 FROM users WHERE user_id=?',(uid,)).fetchone()): await update.message.reply_text('User not found.'); return
    add=ctx.user_data['admin_state']=='bal_uid_add'; ctx.user_data.update(balance_uid=uid,admin_state='bal_amt_add' if add else 'bal_amt_sub'); await update.message.reply_text('Send amount.')
async def balance_amt(update,ctx):
    try:a=float(update.message.text); assert a>0
    except: await update.message.reply_text('Invalid amount.'); return
    uid=ctx.user_data['balance_uid']; add=ctx.user_data['admin_state']=='bal_amt_add'; c=db();
    if add: c.execute('UPDATE users SET balance=balance+? WHERE user_id=?',(a,uid))
    else: c.execute('UPDATE users SET balance=CASE WHEN balance>=? THEN balance-? ELSE 0 END WHERE user_id=?',(a,a,uid))
    c.commit(); c.close(); nb=balance(uid); state_clear(ctx); await update.message.reply_text(msg('success','Balance Updated',f'User: <code>{uid}</code>\nNew balance: <b>{money(nb)} ৳</b>'),parse_mode='HTML',reply_markup=user_kb_admin())
async def user_details_start(update,ctx): ctx.user_data['admin_state']='user_details'; await update.message.reply_text('Send user ID.')
async def user_details(update,ctx):
    try:uid=int(update.message.text)
    except: await update.message.reply_text('Invalid user ID.'); return
    c=db(); r=c.execute('SELECT * FROM users WHERE user_id=?',(uid,)).fetchone(); c.close(); state_clear(ctx)
    if not r: await update.message.reply_text('User not found.',reply_markup=user_kb_admin()); return
    await update.message.reply_text(msg('admin','User Details',f'ID: <code>{uid}</code>\nUsername: @{escape(r["username"] or "-")}\nName: {escape(r["first_name"])}\nBalance: <b>{money(r["balance"])} ৳</b>'),parse_mode='HTML',reply_markup=user_kb_admin())

async def pending(update,ctx):
    c=db(); rows=c.execute('SELECT * FROM deposits WHERE status="pending" ORDER BY id DESC LIMIT 50').fetchall(); c.close()
    if not rows: await update.message.reply_text(msg('payment','Pending Deposits','None.'),parse_mode='HTML',reply_markup=admin_kb()); return
    buttons=[]
    for r in rows:
        buttons.append([InlineKeyboardButton(f'Approve #{r["id"]}',callback_data=f'approve:{r["id"]}'),InlineKeyboardButton(f'Reject #{r["id"]}',callback_data=f'reject:{r["id"]}')])
    body='\n\n'.join([f'ID <b>{r["id"]}</b> | User <code>{r["user_id"]}</code> | <b>{money(r["amount"])} ৳</b> | TRX <code>{escape(r["trx_id"])}</code>' for r in rows])
    await update.message.reply_text(msg('payment','Pending Deposits',body),parse_mode='HTML',reply_markup=InlineKeyboardMarkup(buttons))
async def deposit_action(update,ctx):
    q=update.callback_query; await q.answer(); action,did=q.data.split(':'); did=int(did); c=db(); r=c.execute('SELECT * FROM deposits WHERE id=? AND status="pending"',(did,)).fetchone()
    if not r: c.close(); await q.message.reply_text(msg('error','Already Processed','Deposit is no longer pending.'),parse_mode='HTML'); return
    if action=='approve':
        c.execute('UPDATE deposits SET status="approved" WHERE id=?',(did,)); c.execute('UPDATE users SET balance=balance+? WHERE user_id=?',(r['amount'],r['user_id'])); c.commit(); new=balance(r['user_id']); await ctx.bot.send_message(r['user_id'],msg('success','Deposit Approved',f'Added: <b>{money(r["amount"])} ৳</b>\nBalance: <b>{money(new)} ৳</b>'),parse_mode='HTML')
    else: c.execute('UPDATE deposits SET status="rejected" WHERE id=?',(did,)); c.commit()
    c.close(); await q.message.reply_text(msg('success','Done','Deposit status updated.'),parse_mode='HTML')

async def broadcast(update,ctx): ctx.user_data['admin_state']='broadcast'; await update.message.reply_text('Send broadcast text.',reply_markup=back_kb())
async def broadcast_send(update,ctx):
    text=update.message.text; c=db(); ids=[r['user_id'] for r in c.execute('SELECT user_id FROM users').fetchall()]; c.close(); ok=0
    for uid in ids:
        try: await ctx.bot.send_message(uid,text); ok+=1
        except: pass
        await asyncio.sleep(.03)
    state_clear(ctx); await update.message.reply_text(msg('success','Broadcast Complete',f'Sent: <b>{ok}</b>'),parse_mode='HTML',reply_markup=admin_kb())

async def handle(update,ctx):
    if not update.message or not update.effective_user: return
    ensure_user(update.effective_user); uid=update.effective_user.id; text=(update.message.text or '').strip()
    if not is_admin(uid) and not await guard(update,ctx): return
    if text in ('Back','🔙 Return'):
        state_clear(ctx); await update.message.reply_text(msg('home','Main Menu',''),parse_mode='HTML',reply_markup=user_kb(uid)); return
    # admin state router
    st=ctx.user_data.get('admin_state')
    if is_admin(uid) and st:
        mapping={'add_name':add_product_name,'add_price':add_product_price,'add_desc':add_product_desc,'delete_product':delete_product,'edit_id':edit_product_id,'edit_name':edit_product_name,'edit_price':edit_product_price,'edit_desc':edit_product_desc,'stock_product':stock_product,'fj_add':force_add_save,'fj_remove':force_remove_save,'fj_toggle':force_toggle_save,'pc_name':pc_name,'pc_details':pc_details,'pc_min':pc_min,'pc_max':pc_max,'pe_id':pe_id,'pe_details':pe_details,'pe_min':pe_min,'pe_max':pe_max,'pd_id':payment_delete_save,'pt_id':payment_toggle_save,'pl_id':payment_limits_id,'pl_min':payment_limits_min,'pl_max':payment_limits_max,'user_details':user_details,'bal_uid_add':balance_uid,'bal_uid_sub':balance_uid,'bal_amt_add':balance_amt,'bal_amt_sub':balance_amt,'broadcast':broadcast_send}
        if st in mapping: await mapping[st](update,ctx); return
    if st=='deposit_category':
        rid=ctx.user_data.get('deposit_categories',{}).get(text)
        if rid: await payment_category(update,ctx,rid); return
    if st=='deposit_amount': await deposit_amount(update,ctx,text); return
    if st=='deposit_trx': await deposit_trx(update,ctx,text); return
    if ctx.user_data.get('buy_product_id'):
        try:
            if await buy_quantity(update,ctx,int(text)): return
        except: pass
    pb=ctx.user_data.get('product_buttons',{}).get(text)
    if pb: await show_product(update,ctx,pb); return
    if text=='Buy Product': await show_buy(update,ctx); return
    if text=='Deposit Money': await deposit_start(update,ctx); return
    if text=='My Balance': await update.message.reply_text(msg('money','My Balance',f'<b>{money(balance(uid))} ৳</b>'),parse_mode='HTML',reply_markup=user_kb(uid)); return
    if text=='Price List':
        body='\n\n'.join([f'<b>{escape(p["name"])}</b>\nPrice: {money(p["price"])} ৳\nStock: {p["stock_count"]}' for p in products()]) or 'No products.'; await update.message.reply_text(msg('product','Price List',body),parse_mode='HTML',reply_markup=user_kb(uid)); return
    if text=='Temp Number': await update.message.reply_text(msg('product','Temp Number','Open the service:'),parse_mode='HTML',reply_markup=inline_link('Open Temp Number',TEMP_NUMBER_URL)); return
    if text=='Mail Inbox': await update.message.reply_text(msg('product','Mail Inbox','Open the mailbox service:'),parse_mode='HTML',reply_markup=inline_link('Open Mail Inbox',MAIL_INBOX_URL)); return
    if text=='Admin Panel' and is_admin(uid): await admin_panel(update,ctx); return
    if not is_admin(uid): await update.message.reply_text('Use the menu buttons.',reply_markup=user_kb(uid)); return
    # admin menus
    actions={'Product Management':product_management,'Force Join':force_panel,'Payment Management':payment_panel,'User Management':lambda u,c:update.message.reply_text(msg('admin','User Management',''),parse_mode='HTML',reply_markup=user_kb_admin()),'Pending Deposits':pending,'Broadcast':broadcast,'Add Product':add_product_start,'Edit Product':edit_product_start,'Delete Product':delete_product_start,'Products':admin_products,'Add Stock':add_stock_start,'Add Force Join':force_add,'Remove Force Join':force_remove,'Force Join List':force_list,'Force Join Active/Disable':force_toggle,'Add Category':add_cat,'Edit Category':payment_edit,'Delete Category':payment_delete,'Payment Categories':payment_list,'Payment Enable/Disable':payment_toggle,'Set Deposit Min/Max':payment_limits,'Users':users,'Add Balance':lambda u,c:balance_change_start(u,c,True),'Deduct Balance':lambda u,c:balance_change_start(u,c,False),'User Details':user_details_start}
    if text in actions: await actions[text](update,ctx); return
    await update.message.reply_text('Use the menu buttons.',reply_markup=admin_kb())

async def help_cmd(update,ctx): await update.message.reply_text('Use /start to open the menu.')

def main():
    if BOT_TOKEN=='PASTE_YOUR_BOT_TOKEN_HERE': raise RuntimeError('Set BOT_TOKEN environment variable.')
    init_db(); app=ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler('start',start)); app.add_handler(CommandHandler('help',help_cmd))
    app.add_handler(MessageHandler(filters.Document.ALL,handle))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND,handle))
    from telegram.ext import CallbackQueryHandler
    app.add_handler(CallbackQueryHandler(check_join_cb,pattern=r'^check_join$'))
    app.add_handler(CallbackQueryHandler(deposit_action,pattern=r'^(approve|reject):\d+$'))
    app.run_polling(drop_pending_updates=True)

if __name__=='__main__': main()
